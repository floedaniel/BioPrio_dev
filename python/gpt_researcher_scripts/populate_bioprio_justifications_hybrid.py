"""
BioPRIO Database Justification Populator (Hybrid Research Version)

Adapted from FinnPRIO for terrestrial invertebrates. Uses HYBRID research mode
that combines web search with local PDF documents for each species.

Key features:
- HYBRID RESEARCH: Combines web search with local PDF documents
- Local docs loaded from Species/{GBIF_KEY}_{Scientific_Name}/ folder
- Falls back to web-only if no local docs found
- Copies entire database (preserves complete structure)
- Appends AI justifications to answers table
- Handles pathway questions for EACH selected pathway
- Clean plain text output (no markdown)
- Question-specific instructions
- Domain exclusions
- Full cost tracking with Excel export
"""

import os
import asyncio
import sqlite3
import shutil
import stat
import time
from pathlib import Path
from gpt_researcher import GPTResearcher
from gpt_researcher.utils.enum import Tone
from datetime import datetime
from typing import Dict, List, Tuple, Optional
from dataclasses import dataclass, field
import re

# Import instructions loader (auto-generates JSON from Rmd if needed)
from bioprio_instructions_loader import build_justification_prompt

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️  openpyxl not installed. Excel export disabled. Install with: pip install openpyxl")

# =============================================================================
# CONFIGURATION
# =============================================================================

#  IMPORTANT: READ BEFORE RUNNING
#  THIS SCRIPT CREATES A NEW COPY OF YOUR DATABASE EACH TIME IT RUNS!
#  Using original database again will lose all AI work!

# Skip Existing Justifications
#  True  = Skip questions that already have justification text (recommended for re-runs)
#  False = Append new AI text to existing justifications
# NOTE: Applies to ALL questions (regular + pathway). If you add a new pathway,
#       those new pathway questions have no existing justification, so they get filled anyway.
SKIP_EXISTING_JUSTIFICATION = True

# DATABASE PATH - UPDATE THIS IF YOU ADDED PATHWAYS
DEFAULT_DB_PATH = r"C:\Users\dafl\OneDrive - Folkehelseinstituttet\FinnPrio\BioiPRIO_development\databases\ants\ants_High.db"

# Alternative: path to update already existing AI-enhanced database
# DEFAULT_DB_PATH = r"C:\Users\dafl\OneDrive - Folkehelseinstituttet\FinnPrio\BioiPRIO_development\databases\ant_test\clean_ants_ai_enhanced_19_02_2026.db"

# Output directory (new copy will be created here)
DEFAULT_OUTPUT_DIR = r"C:\Users\dafl\OneDrive - Folkehelseinstituttet\FinnPrio\BioiPRIO_development\databases\ants_ai"

# Filter by species identifiers (empty list = process all species)
# Supports: EPPO codes, scientific names, or GBIF taxon keys ["1315155", "1317433"]
SPECIES_FILTER = [ ]

# Filter by question code (None = process all questions)
# Example: QUESTION_FILTER = "EST2"  # Only process EST2
# Pathway questions: "ENT2A", "ENT2B", "ENT3", "ENT4"
QUESTION_FILTER = None

# =============================================================================
# HYBRID RESEARCH - LOCAL DOCUMENTS CONFIGURATION
# =============================================================================

# Base path where species folders with PDFs are stored
SPECIES_DOCS_BASE_PATH = r"C:\Users\dafl\OneDrive - Folkehelseinstituttet\Prosjektdata - Dokumenter\VKM Data\27.02.2025_maur_forprosjekt_biologisk_mangfold\data\species"

# Temp folder name for GPT Researcher local docs (created in script directory)
TEMP_DOCS_FOLDER = "my-docs"

# File extensions to include in hybrid research
DOCUMENT_EXTENSIONS = {".pdf", ".txt", ".docx", ".doc"}

# =============================================================================
# API Keys - Read from files
OPENAI_API_KEY_FILE = r"C:\Users\dafl\Desktop\API keys\tore_vkm_openai.txt"
TAVILY_API_KEY_FILE = r"C:\Users\dafl\Desktop\API keys\Tavily_key.txt"

# Load API keys from files
def load_api_key(file_path: str) -> str:
    """Load API key from file, stripping whitespace"""
    try:
        with open(file_path, 'r') as f:
            return f.read().strip()
    except FileNotFoundError:
        print(f"⚠️  Warning: API key file not found: {file_path}")
        return ""

os.environ['OPENAI_API_KEY'] = load_api_key(OPENAI_API_KEY_FILE)
os.environ['TAVILY_API_KEY'] = load_api_key(TAVILY_API_KEY_FILE)
# =============================================================================

# GPT Researcher Configuration
os.environ.update({
    "TEMPERATURE": "0.1",
    # GPT Researcher requires the '<provider>:<model>' format for these vars.
    "FAST_LLM": "openai:gpt-4o-mini",
    "SMART_LLM": "openai:gpt-4o",
    "STRATEGIC_LLM": "openai:gpt-4o",
    "MAX_TOKENS": "8000",
    "MAX_SEARCH_RESULTS_PER_QUERY": "15",
    "MAX_URLS_TO_SCRAPE": "20",
    "TOTAL_WORDS": "1000",
    "MAX_ITERATIONS": "8",
    "SIMILARITY_THRESHOLD": "0.38",
    "REPORT_FORMAT": "apa",
})

# Excluded domains
EXCLUDED_DOMAINS = [
    "grokipedia.com",
    "wikipedia.org",
]

# Cost Tracking Configuration
ENABLE_COST_TRACKING = True  # Set to False to disable Excel cost report
COST_REPORT_FILENAME = "cost_report"  # Will append timestamp

# OpenAI pricing (per 1M tokens) - Update as needed
# Prices as of 2024 - check OpenAI pricing page for current rates
OPENAI_PRICING = {
    "gpt-4o-mini": {"input": 0.15, "output": 0.60},
    "gpt-4o": {"input": 2.50, "output": 10.00},
    "gpt-4-turbo": {"input": 10.00, "output": 30.00},
    "gpt-3.5-turbo": {"input": 0.50, "output": 1.50},
}

# Tavily pricing (per search) - approximate
TAVILY_PRICE_PER_SEARCH = 0.01  # Approximate cost per search

# =============================================================================
# COST TRACKING DATA STRUCTURES
# =============================================================================

@dataclass
class QuestionMetrics:
    """Metrics for a single question research call."""
    species_name: str
    question_code: str
    question_text: str
    pathway_name: Optional[str] = None
    start_time: Optional[datetime] = None
    end_time: Optional[datetime] = None
    duration_seconds: float = 0.0
    input_tokens: int = 0
    output_tokens: int = 0
    total_tokens: int = 0
    search_count: int = 0
    estimated_llm_cost: float = 0.0
    estimated_search_cost: float = 0.0
    estimated_total_cost: float = 0.0
    output_chars: int = 0
    status: str = "pending"  # pending, success, error, skipped
    error_message: str = ""


@dataclass
class SpeciesMetrics:
    """Aggregated metrics for a species."""
    species_name: str
    eppo_code: str
    assessment_id: int
    start_time: Optional[datetime] = None
    end_time: Optional[datetime] = None
    total_duration_seconds: float = 0.0
    questions_processed: int = 0
    questions_skipped: int = 0
    questions_failed: int = 0
    total_input_tokens: int = 0
    total_output_tokens: int = 0
    total_tokens: int = 0
    total_searches: int = 0
    estimated_llm_cost: float = 0.0
    estimated_search_cost: float = 0.0
    estimated_total_cost: float = 0.0
    question_metrics: List[QuestionMetrics] = field(default_factory=list)


class CostTracker:
    """Tracks costs and metrics across the entire run."""

    def __init__(self, model_name: str = "gpt-4o-mini"):
        self.model_name = model_name
        self.species_metrics: List[SpeciesMetrics] = []
        self.current_species: Optional[SpeciesMetrics] = None
        self.run_start_time: datetime = datetime.now()
        self.run_end_time: Optional[datetime] = None

        # Get pricing for model
        self.input_price_per_1m = OPENAI_PRICING.get(model_name, {}).get("input", 0.15)
        self.output_price_per_1m = OPENAI_PRICING.get(model_name, {}).get("output", 0.60)

    def start_species(self, species_name: str, eppo_code: str, assessment_id: int):
        """Start tracking a new species."""
        self.current_species = SpeciesMetrics(
            species_name=species_name,
            eppo_code=eppo_code,
            assessment_id=assessment_id,
            start_time=datetime.now()
        )

    def end_species(self):
        """End tracking current species and add to list."""
        if self.current_species:
            self.current_species.end_time = datetime.now()
            self.current_species.total_duration_seconds = (
                self.current_species.end_time - self.current_species.start_time
            ).total_seconds()

            # Aggregate question metrics
            for qm in self.current_species.question_metrics:
                self.current_species.total_input_tokens += qm.input_tokens
                self.current_species.total_output_tokens += qm.output_tokens
                self.current_species.total_tokens += qm.total_tokens
                self.current_species.total_searches += qm.search_count
                self.current_species.estimated_llm_cost += qm.estimated_llm_cost
                self.current_species.estimated_search_cost += qm.estimated_search_cost
                self.current_species.estimated_total_cost += qm.estimated_total_cost

            self.species_metrics.append(self.current_species)
            self.current_species = None

    def record_question(self, metrics: QuestionMetrics):
        """Record metrics for a question."""
        if self.current_species:
            # Use real LLM cost from researcher.get_costs() if available;
            # otherwise fall back to token × price estimate.
            if metrics.estimated_llm_cost == 0:
                metrics.estimated_llm_cost = (
                    (metrics.input_tokens * self.input_price_per_1m / 1_000_000) +
                    (metrics.output_tokens * self.output_price_per_1m / 1_000_000)
                )
            metrics.estimated_search_cost = metrics.search_count * TAVILY_PRICE_PER_SEARCH
            metrics.estimated_total_cost = metrics.estimated_llm_cost + metrics.estimated_search_cost

            self.current_species.question_metrics.append(metrics)

            if metrics.status == "success":
                self.current_species.questions_processed += 1
            elif metrics.status == "skipped":
                self.current_species.questions_skipped += 1
            elif metrics.status == "error":
                self.current_species.questions_failed += 1

    def finalize(self):
        """Finalize the run and calculate totals."""
        self.run_end_time = datetime.now()

    def get_totals(self) -> Dict:
        """Get total metrics across all species."""
        totals = {
            "total_species": len(self.species_metrics),
            "total_duration_seconds": sum(s.total_duration_seconds for s in self.species_metrics),
            "total_questions_processed": sum(s.questions_processed for s in self.species_metrics),
            "total_questions_skipped": sum(s.questions_skipped for s in self.species_metrics),
            "total_questions_failed": sum(s.questions_failed for s in self.species_metrics),
            "total_input_tokens": sum(s.total_input_tokens for s in self.species_metrics),
            "total_output_tokens": sum(s.total_output_tokens for s in self.species_metrics),
            "total_tokens": sum(s.total_tokens for s in self.species_metrics),
            "total_searches": sum(s.total_searches for s in self.species_metrics),
            "estimated_llm_cost": sum(s.estimated_llm_cost for s in self.species_metrics),
            "estimated_search_cost": sum(s.estimated_search_cost for s in self.species_metrics),
            "estimated_total_cost": sum(s.estimated_total_cost for s in self.species_metrics),
        }
        return totals

    def export_to_excel(self, output_dir: str, filename: str = None):
        """Export metrics to Excel file."""
        if not EXCEL_AVAILABLE:
            print("⚠️  Cannot export to Excel: openpyxl not installed")
            return None

        if not self.species_metrics:
            print("⚠️  No metrics to export")
            return None

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = filename or COST_REPORT_FILENAME
        filepath = Path(output_dir) / f"{filename}_{timestamp}.xlsx"

        wb = openpyxl.Workbook()

        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        money_format = '"$"#,##0.0000'
        number_format = '#,##0'
        time_format = '#,##0.00'

        # === Sheet 1: Summary ===
        ws_summary = wb.active
        ws_summary.title = "Summary"

        totals = self.get_totals()
        run_duration = (self.run_end_time - self.run_start_time).total_seconds() if self.run_end_time else 0

        summary_data = [
            ["BioPRIO Cost Report", ""],
            ["", ""],
            ["Run Information", ""],
            ["Start Time", self.run_start_time.strftime("%Y-%m-%d %H:%M:%S")],
            ["End Time", self.run_end_time.strftime("%Y-%m-%d %H:%M:%S") if self.run_end_time else "N/A"],
            ["Total Duration", f"{run_duration/60:.1f} minutes"],
            ["Model", self.model_name],
            ["", ""],
            ["Processing Summary", ""],
            ["Species Processed", totals["total_species"]],
            ["Questions Processed", totals["total_questions_processed"]],
            ["Questions Skipped", totals["total_questions_skipped"]],
            ["Questions Failed", totals["total_questions_failed"]],
            ["", ""],
            ["Token Usage", ""],
            ["Input Tokens", totals["total_input_tokens"]],
            ["Output Tokens", totals["total_output_tokens"]],
            ["Total Tokens", totals["total_tokens"]],
            ["Total Searches", totals["total_searches"]],
            ["", ""],
            ["Cost Estimates", ""],
            ["LLM Cost", f"${totals['estimated_llm_cost']:.4f}"],
            ["Search Cost", f"${totals['estimated_search_cost']:.4f}"],
            ["Total Cost", f"${totals['estimated_total_cost']:.4f}"],
            ["", ""],
            ["Average per Species", ""],
            ["Avg Cost/Species", f"${totals['estimated_total_cost']/max(1,totals['total_species']):.4f}"],
            ["Avg Time/Species", f"{totals['total_duration_seconds']/max(1,totals['total_species'])/60:.1f} min"],
        ]

        for row_idx, row in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = Font(bold=True, size=14)
                elif value in ["Run Information", "Processing Summary", "Token Usage",
                              "Cost Estimates", "Average per Species"]:
                    cell.font = Font(bold=True)

        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 25

        # === Sheet 2: Species Summary ===
        ws_species = wb.create_sheet("Per Species")

        species_headers = [
            "Species Name", "EPPO Code", "Assessment ID", "Duration (min)",
            "Questions Done", "Skipped", "Failed",
            "Input Tokens", "Output Tokens", "Total Tokens", "Searches",
            "LLM Cost ($)", "Search Cost ($)", "Total Cost ($)"
        ]

        for col_idx, header in enumerate(species_headers, 1):
            cell = ws_species.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        for row_idx, sm in enumerate(self.species_metrics, 2):
            data = [
                sm.species_name,
                sm.eppo_code,
                sm.assessment_id,
                sm.total_duration_seconds / 60,
                sm.questions_processed,
                sm.questions_skipped,
                sm.questions_failed,
                sm.total_input_tokens,
                sm.total_output_tokens,
                sm.total_tokens,
                sm.total_searches,
                sm.estimated_llm_cost,
                sm.estimated_search_cost,
                sm.estimated_total_cost,
            ]
            for col_idx, value in enumerate(data, 1):
                cell = ws_species.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                if col_idx == 4:
                    cell.number_format = time_format
                elif col_idx in [8, 9, 10, 11]:
                    cell.number_format = number_format
                elif col_idx in [12, 13, 14]:
                    cell.number_format = money_format

        # Auto-width columns
        for col in ws_species.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws_species.column_dimensions[col[0].column_letter].width = min(max_length + 2, 20)

        # === Sheet 3: Question Details ===
        ws_questions = wb.create_sheet("Question Details")

        question_headers = [
            "Species", "Question Code", "Pathway", "Status",
            "Duration (s)", "Input Tokens", "Output Tokens", "Searches",
            "LLM Cost ($)", "Search Cost ($)", "Total Cost ($)",
            "Output Chars", "Error"
        ]

        for col_idx, header in enumerate(question_headers, 1):
            cell = ws_questions.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        row_idx = 2
        for sm in self.species_metrics:
            for qm in sm.question_metrics:
                data = [
                    sm.species_name,
                    qm.question_code,
                    qm.pathway_name or "",
                    qm.status,
                    qm.duration_seconds,
                    qm.input_tokens,
                    qm.output_tokens,
                    qm.search_count,
                    qm.estimated_llm_cost,
                    qm.estimated_search_cost,
                    qm.estimated_total_cost,
                    qm.output_chars,
                    qm.error_message,
                ]
                for col_idx, value in enumerate(data, 1):
                    cell = ws_questions.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    if col_idx == 5:
                        cell.number_format = time_format
                    elif col_idx in [6, 7, 8, 12]:
                        cell.number_format = number_format
                    elif col_idx in [9, 10, 11]:
                        cell.number_format = money_format
                row_idx += 1

        # Auto-width columns
        for col in ws_questions.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws_questions.column_dimensions[col[0].column_letter].width = min(max_length + 2, 25)

        wb.save(filepath)
        print(f"\n📊 Cost report saved: {filepath}")
        return str(filepath)


# Global cost tracker instance
cost_tracker: Optional[CostTracker] = None

# =============================================================================
# TEXT CLEANING FUNCTIONS
# =============================================================================

def clean_markdown_formatting(text: str) -> str:
    """Remove markdown formatting and clean up AI-generated text."""

    # Remove markdown headings
    text = re.sub(r'^#+\s+.*$', '', text, flags=re.MULTILINE)

    # Remove bold/italic
    text = re.sub(r'\*\*([^*]+)\*\*', r'\1', text)
    text = re.sub(r'\*([^*]+)\*', r'\1', text)
    text = re.sub(r'__([^_]+)__', r'\1', text)
    text = re.sub(r'_([^_]+)_', r'\1', text)

    # Remove markdown links
    text = re.sub(r'\[([^]]+)]\([^)]+\)', r'\1', text)

    # Remove markdown tables
    text = re.sub(r'^\s*\|[^\n]+\|\s*$', '', text, flags=re.MULTILINE)
    text = re.sub(r'^\s*\|[\s\-:|]+\|\s*$', '', text, flags=re.MULTILINE)

    # Remove bullet points
    text = re.sub(r'^\s*[-*+]\s+', '', text, flags=re.MULTILINE)
    text = re.sub(r'^\s*\d+\.\s+', '', text, flags=re.MULTILINE)

    # Remove code blocks
    text = re.sub(r'```[\s\S]*?```', '', text)
    text = re.sub(r'`([^`]+)`', r'\1', text)

    # Remove horizontal rules
    text = re.sub(r'^[-*_]{3,}$', '', text, flags=re.MULTILINE)

    # Remove separator phrases
    separators = [
        r'---\s*\*\*AI-Generated.*?\*\*\s*---',
        r'\*\*AI-Generated.*?\*\*',
        r'---\s*AI-Generated.*?---',
        r'AI-Generated Supplementary Information.*?\n',
        r'\(GPT Researcher\)',
    ]
    for pattern in separators:
        text = re.sub(pattern, '', text, flags=re.MULTILINE | re.DOTALL)

    # Remove common AI introduction phrases
    intro_patterns = [
        r'^.*?[Ii]ntroduction.*?$',
        r'^.*?[Ss]ummary.*?$',
        r'^.*?[Oo]verview.*?$',
        r'^This report.*?$',
    ]
    for pattern in intro_patterns:
        text = re.sub(pattern, '', text, flags=re.MULTILINE)

    # Clean up whitespace
    text = re.sub(r'\n{3,}', '\n\n', text)
    text = re.sub(r'[ \t]+', ' ', text)
    text = re.sub(r'\n ', '\n', text)
    text = re.sub(r' \n', '\n', text)
    text = text.strip()

    return text

# =============================================================================
# LOCAL DOCUMENT FUNCTIONS
# =============================================================================

def force_rmtree(path: Path) -> None:
    """Remove a directory tree, forcing read-only files to be deleted (Windows-safe)."""
    def _on_error(func, target, _):
        os.chmod(target, stat.S_IWRITE)
        func(target)
    shutil.rmtree(path, onerror=_on_error)


def find_species_docs_folder(gbif_key: str, scientific_name: str) -> Optional[Path]:
    """Find the species folder matching {GBIF_KEY}_{Scientific_Name} pattern.

    Args:
        gbif_key: GBIF taxon key (e.g., "11700741")
        scientific_name: Species name (e.g., "Lasius aphidicola")

    Returns:
        Path to folder if found, None otherwise.
    """
    if not gbif_key or not scientific_name:
        return None

    base_path = Path(SPECIES_DOCS_BASE_PATH)
    if not base_path.exists():
        print(f"  ⚠️  Species docs base path not found: {SPECIES_DOCS_BASE_PATH}")
        return None

    # Build expected folder name: {gbif_key}_{scientific_name_with_underscores}
    safe_name = scientific_name.replace(" ", "_")
    expected_folder = f"{gbif_key}_{safe_name}"

    # Try exact match first
    exact_path = base_path / expected_folder
    if exact_path.exists():
        return exact_path

    # Try case-insensitive search
    for folder in base_path.iterdir():
        if folder.is_dir() and folder.name.lower() == expected_folder.lower():
            return folder

    # Try partial match (folder starts with GBIF key)
    for folder in base_path.iterdir():
        if folder.is_dir() and folder.name.startswith(f"{gbif_key}_"):
            return folder

    return None


def copy_species_docs_to_temp(gbif_key: str, scientific_name: str) -> bool:
    """Copy all documents from species folder to temp my-docs folder.

    GPT Researcher's hybrid mode reads from a "my-docs" folder in the script directory.

    Args:
        gbif_key: GBIF taxon key
        scientific_name: Species scientific name

    Returns:
        True if docs were copied (use hybrid mode), False otherwise (use web-only).
    """
    # Get script directory for temp folder location
    script_dir = Path(__file__).parent
    temp_path = script_dir / TEMP_DOCS_FOLDER

    # Clear existing temp folder
    if temp_path.exists():
        force_rmtree(temp_path)
    temp_path.mkdir(parents=True, exist_ok=True)

    # Find species folder
    species_path = find_species_docs_folder(gbif_key, scientific_name)
    if not species_path:
        print(f"  ⚠️  No local documents folder found for {gbif_key}_{scientific_name}")
        return False

    print(f"  📂 Found species docs: {species_path.name}")

    # Recursively find all matching documents
    docs_copied = 0
    for ext in DOCUMENT_EXTENSIONS:
        for doc_file in species_path.rglob(f"*{ext}"):
            if doc_file.is_file():
                # Copy to flat structure with unique names (avoid collisions)
                dest_name = f"{docs_copied:04d}_{doc_file.name}"
                dest_path = temp_path / dest_name
                try:
                    shutil.copy2(doc_file, dest_path)
                    docs_copied += 1
                except Exception as e:
                    print(f"  ⚠️  Failed to copy {doc_file.name}: {e}")

    if docs_copied > 0:
        print(f"  📚 Copied {docs_copied} documents to temp folder for hybrid research")
        return True
    else:
        print(f"  ⚠️  No documents found in {species_path}")
        return False


def cleanup_temp_docs():
    """Remove temp my-docs folder."""
    script_dir = Path(__file__).parent
    temp_path = script_dir / TEMP_DOCS_FOLDER
    if temp_path.exists():
        try:
            force_rmtree(temp_path)
            print("🧹 Cleaned up temp documents folder")
        except Exception as e:
            print(f"⚠️  Failed to cleanup temp folder: {e}")


# =============================================================================
# DATABASE FUNCTIONS - GENERAL
# =============================================================================

def copy_database(source_path: str, output_dir: str) -> str:
    """Copy entire source database to new location."""
    # Get original database name without extension
    source_file = Path(source_path)
    original_name = source_file.stem  # filename without .db

    # Create timestamp in DD_MM_YYYY format
    timestamp = datetime.now().strftime("%d_%m_%Y")

    # Check if source already has _ai_enhanced_ pattern - extract base name
    if "_ai_enhanced_" in original_name:
        base_name = original_name.split("_ai_enhanced_")[0]
    else:
        base_name = original_name

    # New name: base_name_ai_enhanced_DD_MM_YYYY.db
    output_name = f"{base_name}_ai_enhanced_{timestamp}.db"
    output_path = Path(output_dir) / output_name

    Path(output_dir).mkdir(parents=True, exist_ok=True)

    # Check if source and destination are the same file (re-run on same day)
    if source_file.resolve() == output_path.resolve():
        print(f"\n📋 Using existing database (same-day re-run)...")
        print(f"   Path: {source_path}")
        print(f"✅ Working on existing file ({output_path.stat().st_size / 1024:.1f} KB)")
        return str(output_path)

    print(f"\n📋 Copying database...")
    print(f"   From: {source_path}")
    print(f"   To:   {output_path}")

    shutil.copy2(source_path, output_path)

    if output_path.exists():
        print(f"✅ Database copied successfully ({output_path.stat().st_size / 1024:.1f} KB)")
    else:
        raise FileNotFoundError(f"Failed to copy database to {output_path}")

    return str(output_path)

def get_all_assessment_ids(db_path: str, species_filter: List[str] = None) -> List[int]:
    """Get all assessment IDs, optionally filtered by species identifiers."""
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    if species_filter:
        # Filter by EPPO codes, scientific names, or GBIF keys (case-insensitive)
        placeholders = ','.join(['?' for _ in species_filter])
        upper_filters = [f.upper() for f in species_filter]
        original_filters = species_filter

        cursor.execute(f"""
            SELECT a.idAssessment
            FROM assessments a
            JOIN pests p ON a.idPest = p.idPest
            WHERE UPPER(p.eppoCode) IN ({placeholders})
               OR UPPER(p.scientificName) IN ({placeholders})
               OR p.gbifTaxonKey IN ({placeholders})
            ORDER BY a.idAssessment
        """, upper_filters + upper_filters + original_filters)
    else:
        cursor.execute("""
            SELECT idAssessment
            FROM assessments
            ORDER BY idAssessment
        """)

    ids = [row[0] for row in cursor.fetchall()]
    conn.close()
    return ids


def get_species_identifiers_for_assessments(db_path: str, assessment_ids: List[int]) -> List[str]:
    """Get species identifiers for assessments."""
    if not assessment_ids:
        return []
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    placeholders = ','.join(['?' for _ in assessment_ids])
    cursor.execute(f"""
        SELECT DISTINCT
            COALESCE(NULLIF(p.eppoCode, ''), p.scientificName) as identifier
        FROM assessments a
        JOIN pests p ON a.idPest = p.idPest
        WHERE a.idAssessment IN ({placeholders})
    """, assessment_ids)
    identifiers = [row[0] for row in cursor.fetchall() if row[0]]
    conn.close()
    return identifiers


def get_assessment_info(db_path: str, assessment_id: int) -> Optional[Dict]:
    """Get assessment details including species and regular questions.

    Creates answer rows if they don't exist (needed for script-populated assessments).
    """
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    # Get assessment
    cursor.execute("""
        SELECT a.idAssessment, a.idPest, p.scientificName, p.eppoCode, p.gbifTaxonKey
        FROM assessments a
        JOIN pests p ON a.idPest = p.idPest
        WHERE a.idAssessment = ?
    """, (assessment_id,))

    result = cursor.fetchone()

    if not result:
        conn.close()
        return None

    assessment_id, pest_id, species_name, eppo_code, gbif_key = result

    # Get ALL questions (not dependent on existing answers)
    cursor.execute("""
        SELECT idQuestion, "group", number, subgroup, question, info
        FROM questions
        ORDER BY idQuestion
    """)

    questions = cursor.fetchall()

    # For each question, get or create answer row
    answers = []
    created_count = 0

    for q in questions:
        id_question, grp, num, subgrp, text, info = q

        # Check if answer exists
        cursor.execute("""
            SELECT idAnswer, justification FROM answers
            WHERE idAssessment = ? AND idQuestion = ?
        """, (assessment_id, id_question))

        answer_row = cursor.fetchone()

        if answer_row:
            id_answer, justification = answer_row
        else:
            # Create answer row (same as app does dynamically)
            # Columns: idAssessment, idQuestion, min, likely, max, justification
            cursor.execute("""
                INSERT INTO answers (idAssessment, idQuestion, min, likely, max, justification)
                VALUES (?, ?, '', '', '', '')
            """, (assessment_id, id_question))
            id_answer = cursor.lastrowid
            justification = ""
            created_count += 1

        code = f"{grp}{num}.{subgrp}" if subgrp else f"{grp}{num}."
        answers.append({
            'idAnswer': id_answer,
            'code': code,
            'text': text,
            'info': info or "",
            'existing_justification': justification or ""
        })

    # Commit any new answer rows
    if created_count > 0:
        conn.commit()
        print(f"  ℹ️  Created {created_count} answer rows for this assessment")

    conn.close()

    return {
        'idAssessment': assessment_id,
        'idPest': pest_id,
        'scientificName': species_name,
        'eppoCode': eppo_code,
        'gbifTaxonKey': gbif_key or "",
        'answers': answers
    }

def update_answer_justification(db_path: str, id_answer: int, justification: str):
    """Update justification in answers table."""
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()

        # Verify table exists
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='answers'")
        if not cursor.fetchone():
            raise Exception(f"Table 'answers' not found in database: {db_path}")

        cursor.execute("UPDATE answers SET justification = ? WHERE idAnswer = ?",
                      (justification, id_answer))

        # Verify row was actually updated
        if cursor.rowcount == 0:
            raise Exception(f"No row found with idAnswer={id_answer}")

        conn.commit()
        conn.close()
    except Exception as e:
        print(f"  ⚠️  Database error in update_answer_justification:")
        print(f"     Database: {db_path}")
        print(f"     Answer ID: {id_answer}")
        print(f"     Error: {e}")
        raise

# =============================================================================
# DATABASE FUNCTIONS - PATHWAYS
# =============================================================================

def get_all_available_pathways(db_path: str) -> List[Dict]:
    """Get all available pathways from the pathways table."""
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    cursor.execute("""
        SELECT idPathway, name, "group"
        FROM pathways
        ORDER BY idPathway
    """)

    pathways = []
    for row in cursor.fetchall():
        id_pathway, name, group = row
        pathways.append({
            'idPathway': id_pathway,
            'name': name,
            'group': group
        })

    conn.close()
    return pathways


def add_pathways_to_assessment(db_path: str, assessment_id: int,
                                pathway_ids: List[int] = None) -> int:
    """Add pathways to an assessment if they don't exist.

    Args:
        db_path: Database path
        assessment_id: Assessment ID
        pathway_ids: List of pathway IDs to add. If None, adds all available pathways.

    Returns:
        Number of pathways added
    """
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    # Get all available pathways if none specified
    if pathway_ids is None:
        cursor.execute("SELECT idPathway FROM pathways ORDER BY idPathway")
        pathway_ids = [row[0] for row in cursor.fetchall()]

    # Get existing pathways for this assessment
    cursor.execute("""
        SELECT idPathway FROM entryPathways WHERE idAssessment = ?
    """, (assessment_id,))
    existing_ids = {row[0] for row in cursor.fetchall()}

    # Add missing pathways
    added = 0
    for pid in pathway_ids:
        if pid not in existing_ids:
            cursor.execute("""
                INSERT INTO entryPathways (idAssessment, idPathway, specification)
                VALUES (?, ?, '')
            """, (assessment_id, pid))
            added += 1

    conn.commit()
    conn.close()
    return added


def get_assessment_pathways(db_path: str, assessment_id: int) -> List[Dict]:
    """Get all selected pathways for an assessment."""
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    cursor.execute("""
        SELECT ep.idEntryPathway, ep.idPathway, p.name, p."group", ep.specification
        FROM entryPathways ep
        JOIN pathways p ON ep.idPathway = p.idPathway
        WHERE ep.idAssessment = ?
        ORDER BY p.idPathway
    """, (assessment_id,))

    pathways = []
    for row in cursor.fetchall():
        id_entry, id_pathway, name, group, spec = row
        pathways.append({
            'idEntryPathway': id_entry,
            'idPathway': id_pathway,
            'name': name,
            'group': group,
            'specification': spec or ""
        })

    conn.close()
    return pathways

def get_pathway_questions(db_path: str) -> List[Dict]:
    """Get all pathway questions (ENT2A, ENT2B, ENT3, ENT4)."""
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    cursor.execute("""
        SELECT idPathQuestion, "group", number, question, info
        FROM pathwayQuestions
        ORDER BY idPathQuestion
    """)

    # Map idPathQuestion to proper codes (A/B suffix not stored in DB)
    # idPathQuestion 1 = ENT2A, 2 = ENT2B, 3 = ENT3, 4 = ENT4
    id_to_code = {
        1: "ENT2A",
        2: "ENT2B",
        3: "ENT3",
        4: "ENT4"
    }

    questions = []
    for row in cursor.fetchall():
        id_q, grp, num, text, info = row
        # Use mapping if available, fallback to group+number
        code = id_to_code.get(id_q, f"{grp}{num}")
        questions.append({
            'idPathQuestion': id_q,
            'code': code,
            'text': text,
            'info': info or ""
        })

    conn.close()
    return questions

def get_existing_pathway_justification(db_path: str, id_entry_pathway: int,
                                       id_path_question: int) -> str:
    """Get existing pathway justification."""
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT justification FROM pathwayAnswers
        WHERE idEntryPathway = ? AND idPathQuestion = ?
    """, (id_entry_pathway, id_path_question))
    result = cursor.fetchone()
    conn.close()
    return result[0] if result and result[0] else ""

def update_pathway_justification(db_path: str, id_entry_pathway: int,
                                 id_path_question: int, justification: str):
    """Update or insert pathway justification.

    Note: The pathwayAnswers table requires min, likely, max columns alongside
    justification. When inserting new rows, we set these to empty strings
    (matching app behavior). Values are populated later by populate_bioprio_values.py.
    """
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()

        # Verify table exists
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='pathwayAnswers'")
        if not cursor.fetchone():
            raise Exception(f"Table 'pathwayAnswers' not found in database: {db_path}")

        # Check if exists
        cursor.execute("""
            SELECT idPathAnswer FROM pathwayAnswers
            WHERE idEntryPathway = ? AND idPathQuestion = ?
        """, (id_entry_pathway, id_path_question))

        result = cursor.fetchone()

        if result:
            cursor.execute("""
                UPDATE pathwayAnswers SET justification = ?
                WHERE idPathAnswer = ?
            """, (justification, result[0]))
            # Verify update succeeded
            if cursor.rowcount == 0:
                raise Exception(f"Update failed for idPathAnswer={result[0]}")
        else:
            # Insert with all required columns: min, likely, max as empty strings
            # This matches the app's expected format (server.R lines 1608-1610)
            cursor.execute("""
                INSERT INTO pathwayAnswers (idEntryPathway, idPathQuestion, min, likely, max, justification)
                VALUES (?, ?, '', '', '', ?)
            """, (id_entry_pathway, id_path_question, justification))
            # Verify insert succeeded
            if cursor.lastrowid == 0:
                raise Exception(f"Insert failed for pathway answer")

        conn.commit()
        conn.close()
    except Exception as e:
        print(f"  ⚠️  Database error in update_pathway_justification:")
        print(f"     Database: {db_path}")
        print(f"     EntryPathway ID: {id_entry_pathway}, PathQuestion ID: {id_path_question}")
        print(f"     Error: {e}")
        raise

# =============================================================================
# QUESTION-SPECIFIC INSTRUCTIONS
# =============================================================================

def get_question_specific_instructions(question_code: str, species_name: str,
                                       pathway_name: str = None) -> str:
    """Get research instructions from Rmd-derived JSON.

    Uses the bioprio_instructions_loader to build prompts from the
    Instructions_BioPrio_assessments.rmd file.

    Falls back to hardcoded instructions for pathway-specific cases
    (intentional introduction special handling).
    """
    # Handle pathway-specific special cases (intentional introduction)
    if pathway_name:
        is_intentional = pathway_name.lower() == "intentional introduction"

        # Special handling for intentional introduction pathway
        if is_intentional:
            if question_code == "ENT2A":
                return f"""
For the pathway "Intentional introduction", transport to the risk assessment area is GUARANTEED by definition.

Intentional introduction means {species_name} is deliberately imported by humans
(e.g., for pet trade, biological control, research, hobby keeping, gardening, etc.).

IMPORTANT: Since this is intentional introduction, the probability of transport is MAXIMUM/CERTAIN.
The species will definitely be transported if someone intends to import it.

Briefly describe:
- Known reasons why {species_name} might be intentionally introduced (pet trade, hobby, research, etc.)
- Evidence of intentional trade or keeping of this species
- Legal status regarding import of this species
"""
            elif question_code == "ENT2B":
                return f"""
For the pathway "Intentional introduction", transport probability remains MAXIMUM even with management measures.

When someone intentionally imports {species_name}, official management measures (inspections, regulations)
are often circumvented or ineffective because:
- The importer actively wants to bring the species in
- Illegal imports bypass official channels entirely
- Small specimens (eggs, juveniles) are easily concealed
- Online trade and postal services are difficult to monitor

IMPORTANT: Since this is intentional introduction, management measures provide minimal reduction.
The probability remains MAXIMUM/VERY HIGH.

Briefly describe:
- How regulations might be circumvented for intentional imports of {species_name}
- Effectiveness (or lack thereof) of border controls for this type of import
- Known cases of illegal or unreported imports of similar species
"""
            elif question_code == "ENT3":
                return f"""
For the pathway "Intentional introduction", trade volume should be assessed as HIGH/MAXIMUM.

Intentional introduction of {species_name} implies active demand and deliberate importation.
Even if official trade statistics are low, the actual volume may be significant due to:
- Unrecorded private imports
- Online purchases from international sellers
- Informal trade networks among hobbyists
- Difficulty tracking small-scale imports

IMPORTANT: For intentional introduction, assume trade volume is HIGH/MAXIMUM because:
- Demand drives supply regardless of official channels
- The internet enables easy international purchasing
- Small organisms are easily shipped in parcels

Briefly describe:
- Known trade or demand for {species_name} (pet trade, hobby, collections, etc.)
- Online availability of this species for purchase
- Size of the hobbyist/collector community interested in this species
"""

    # Use Rmd-based instructions for all other cases
    try:
        return build_justification_prompt(question_code, species_name, pathway_name)
    except KeyError:
        # Fallback if question not found in Rmd
        return f"Answer the question about {species_name} based on available scientific evidence."

# =============================================================================
# GPT RESEARCHER FUNCTIONS
# =============================================================================

def create_research_query(species_name: str, question_code: str, question_text: str,
                          question_info: str = "", pathway_name: str = None) -> str:
    """Create targeted research query.

    Note: question_info from database is IGNORED when Rmd instructions are available,
    as the Rmd provides more accurate and up-to-date guidance.
    """

    # Get specific instructions from Rmd (preferred) or hardcoded fallback
    specific = get_question_specific_instructions(question_code, species_name, pathway_name)

    # Check if we got Rmd instructions (they include "QUESTION" header)
    using_rmd_instructions = specific and "QUESTION" in specific

    # Build query
    pathway_text = f' via the pathway "{pathway_name}"' if pathway_name else ""

    # If using Rmd instructions, the 'specific' prompt already contains everything needed
    if using_rmd_instructions:
        query = f"""
Research the following species for a risk assessment:

SPECIES: {species_name}{pathway_text}

{specific}

CRITICAL: Answer ONLY this specific question. Do NOT include information about other topics.

SCOPE LIMITATION:
- Answer based on documented information for THIS EXACT SPECIES only
- Do NOT extrapolate from related species, congeners, or sister taxa
- Do NOT assume biology, hosts, or behavior based on similar species
- If information is limited for this species, acknowledge it clearly
- "Unknown" or "insufficient data for this species" is a valid answer

RESEARCH REQUIREMENTS:
- Base on peer-reviewed literature, official risk assessments (VKM, Fera, EPPO, EFSA, CABI, USDA, and others)
- Provide specific evidence with citations
- Consider Norwegian/Nordic context (temperate to boreal climate, cold winters)
- Acknowledge uncertainty when evidence is limited
- Keep focused and concise (300-400 words)

INSUFFICIENT INFORMATION:
- If the provided context contains insufficient information to answer the question, explicitly state: "The provided context contains insufficient information to answer the question."
- After stating this, you may provide relevant context that IS available, but clearly note the information gaps

ASSUMPTIONS:
- If making any assumptions, clearly indicate them with phrases like:
  * "Assuming that..."
  * "Based on the assumption that..."
  * "It is assumed that..."
- Clearly distinguish between evidence-based statements and assumptions

OUTPUT FORMAT:
- Write in PLAIN TEXT only - NO markdown (#, ##, **, *, -)
- DO NOT use tables - they are unreadable in plain text
- DO NOT include "Introduction" sections
- Answer the question DIRECTLY
- Use paragraph format with proper punctuation
- Citations in parentheses: (Author, Year)
- Write as continuous text, not lists
- If multiple items, write in sentence form

Provide a clear, evidence-based justification.
"""
    else:
        # Fallback: use old format with database question_info (for special cases like intentional introduction)
        query = f"""
Research the following question about {species_name}{pathway_text}:

QUESTION ({question_code}): {question_text}

{specific if specific else "Focus on answering this specific question."}

CRITICAL: Answer ONLY this specific question. Do NOT include information about other topics.

SCOPE LIMITATION:
- Answer based on documented information for THIS EXACT SPECIES only
- Do NOT extrapolate from related species, congeners, or sister taxa
- Do NOT assume biology, hosts, or behavior based on similar species
- If information is limited for this species, acknowledge it clearly
- "Unknown" or "insufficient data for this species" is a valid answer

RESEARCH REQUIREMENTS:
- Base on peer-reviewed literature, official risk assessments (VKM, Fera, EPPO, EFSA, CABI, USDA, and others)
- Provide specific evidence with citations
- Consider Norwegian/Nordic context (temperate to boreal climate, cold winters)
- Acknowledge uncertainty when evidence is limited
- Keep focused and concise (300-400 words)

INSUFFICIENT INFORMATION:
- If the provided context contains insufficient information to answer the question, explicitly state: "The provided context contains insufficient information to answer the question."
- After stating this, you may provide relevant context that IS available, but clearly note the information gaps

ASSUMPTIONS:
- If making any assumptions, clearly indicate them with phrases like:
  * "Assuming that..."
  * "Based on the assumption that..."
  * "It is assumed that..."
- Clearly distinguish between evidence-based statements and assumptions

{f'ADDITIONAL GUIDANCE: {question_info}' if question_info else ''}

OUTPUT FORMAT:
- Write in PLAIN TEXT only - NO markdown (#, ##, **, *, -)
- DO NOT use tables - they are unreadable in plain text
- DO NOT include "Introduction" sections
- Answer the question DIRECTLY
- Use paragraph format with proper punctuation
- Citations in parentheses: (Author, Year)
- Write as continuous text, not lists
- If multiple items, write in sentence form

Provide a clear, evidence-based justification.
"""

    return query

async def research_justification(species_name: str, question_code: str, question_text: str,
                                 question_info: str = "", pathway_name: str = None,
                                 exclude_domains: List[str] = None,
                                 track_metrics: bool = True,
                                 use_hybrid: bool = False) -> Tuple[str, Optional[QuestionMetrics]]:
    """Research a single justification using GPT Researcher.

    Returns:
        Tuple of (report_text, metrics) where metrics may be None if tracking disabled.
    """
    global cost_tracker

    pathway_text = f" (Pathway: {pathway_name})" if pathway_name else ""
    print(f"\n{'=' * 80}")
    print(f"Researching: {species_name} - {question_code}{pathway_text}")
    print(f"{'=' * 80}\n")

    if exclude_domains:
        print(f"⛔ Excluding: {', '.join(exclude_domains)}")
    print(f"🔬 Research mode: {'hybrid (web + local docs)' if use_hybrid else 'web-only'}")

    # Initialize metrics
    metrics = QuestionMetrics(
        species_name=species_name,
        question_code=question_code,
        question_text=question_text[:100],  # Truncate for storage
        pathway_name=pathway_name,
        start_time=datetime.now()
    ) if track_metrics else None

    query = create_research_query(species_name, question_code, question_text,
                                  question_info, pathway_name)

    # Add domain exclusion
    if exclude_domains:
        domain_filter = f"\n\nIMPORTANT: Do NOT use information from: {', '.join(exclude_domains)}"
        query = query + domain_filter

    report_source = "hybrid" if use_hybrid else "web"

    researcher = GPTResearcher(
        query=query,
        report_type="research_report",
        tone=Tone.Objective,
        report_source=report_source,
    )

    max_retries = 5
    base_wait = 5  # seconds
    report = ""

    try:
        start_time = time.time()
        for attempt in range(max_retries):
            try:
                await researcher.conduct_research()
                report = await researcher.write_report()
                break
            except Exception as rate_exc:
                err_str = str(rate_exc)
                is_rate_limit = "429" in err_str or "rate_limit_exceeded" in err_str
                if is_rate_limit and attempt < max_retries - 1:
                    # Try to parse suggested wait time from error message
                    wait_match = re.search(r'try again in (\d+(?:\.\d+)?)s', err_str)
                    wait_time = float(wait_match.group(1)) if wait_match else base_wait * (2 ** attempt)
                    wait_time = max(wait_time + 2, 5)  # at least 5s buffer
                    print(f"⏳ Rate limit hit (attempt {attempt + 1}/{max_retries}). Waiting {wait_time:.0f}s...")
                    await asyncio.sleep(wait_time)
                    # Re-create researcher for retry
                    researcher = GPTResearcher(
                        query=query,
                        report_type="research_report",
                        tone=Tone.Objective,
                        report_source=report_source,
                    )
                else:
                    raise
        end_time = time.time()

        # Remove excluded domain references
        if exclude_domains:
            for domain in exclude_domains:
                report = re.sub(rf'\[([^]]+)]\([^)]*{re.escape(domain)}[^)]*\)', '', report)
                report = re.sub(rf'https?://\S*{re.escape(domain)}\S*', '', report)

        # Clean markdown
        report = clean_markdown_formatting(report)

        # Update metrics
        if metrics:
            metrics.end_time = datetime.now()
            metrics.duration_seconds = end_time - start_time
            metrics.output_chars = len(report)
            metrics.status = "success"

            # Estimate tokens (approximate: ~4 chars per token for English)
            # Query tokens (input)
            metrics.input_tokens = len(query) // 4
            # Report tokens (output)
            metrics.output_tokens = len(report) // 4
            metrics.total_tokens = metrics.input_tokens + metrics.output_tokens

            # Estimate search count from config (GPT Researcher typically does
            # multiple searches per iteration)
            max_iterations = int(os.environ.get("MAX_ITERATIONS", "5"))
            metrics.search_count = max_iterations * 2  # Conservative estimate

            # Get actual cost from researcher (per gpt-researcher API: get_costs() -> float)
            try:
                actual_cost = researcher.get_costs()
                if actual_cost:
                    metrics.estimated_llm_cost = actual_cost
            except Exception:
                pass  # Fall back to token-based estimate in CostTracker.record_question

            print(f"⏱️  Duration: {metrics.duration_seconds:.1f}s | "
                  f"Tokens: ~{metrics.total_tokens:,} | "
                  f"Output: {metrics.output_chars:,} chars")

        return report, metrics
    except Exception as e:
        print(f"ERROR: {str(e)}")
        if metrics:
            metrics.end_time = datetime.now()
            metrics.duration_seconds = (metrics.end_time - metrics.start_time).total_seconds()
            metrics.status = "error"
            metrics.error_message = str(e)[:200]
        return f"ERROR: {str(e)}", metrics

# =============================================================================
# MAIN WORKFLOW
# =============================================================================

async def process_assessment(db_path: str, assessment_id: int = None,
                             exclude_domains: List[str] = None,
                             limit_questions: int = None,
                             process_pathways: bool = True,
                             skip_existing: bool = True,
                             track_costs: bool = True,
                             add_all_pathways: bool = False,
                             question_filter: str = None):
    """Process assessment: regular questions + pathway questions.

    Args:
        add_all_pathways: If True, automatically add all available pathways to the
            assessment before processing. Useful for comprehensive assessments.
        question_filter: If set, only process this specific question code (e.g., "EST2", "ENT2A").
    """
    global cost_tracker

    print("\n📚 Loading assessment data...")
    assessment_info = get_assessment_info(db_path, assessment_id)

    if not assessment_info:
        print("❌ No assessment found!")
        return

    species_name = assessment_info['scientificName']
    eppo_code = assessment_info['eppoCode']
    answers = assessment_info['answers']
    assessment_id = assessment_info['idAssessment']
    gbif_key = assessment_info.get('gbifTaxonKey', '')

    # Set up local documents for hybrid research
    use_hybrid = copy_species_docs_to_temp(gbif_key, species_name)

    # Filter to specific question if requested
    if question_filter:
        # Strip trailing dots for comparison (codes stored as "EST2." but user enters "EST2")
        filter_code = question_filter.upper().rstrip('.')
        answers = [a for a in answers if a['code'].upper().rstrip('.') == filter_code]
        print(f"🔍 Filtering to question: {filter_code}")
        if not answers:
            print(f"⚠️  No matching regular question found for {filter_code}")

    # Auto-add all pathways if requested
    if add_all_pathways and process_pathways:
        added = add_pathways_to_assessment(db_path, assessment_id)
        if added > 0:
            print(f"➕ Auto-added {added} pathways to assessment")
            # Show available pathways
            all_pathways = get_all_available_pathways(db_path)
            print(f"   Available pathways: {', '.join(p['name'] for p in all_pathways)}")

    # Start tracking this species
    if track_costs and cost_tracker:
        cost_tracker.start_species(species_name, eppo_code or "", assessment_id)

    if limit_questions:
        answers = answers[:limit_questions]
        print(f"⚠️  Limited to {limit_questions} questions")

    print(f"\n📊 Assessment: {species_name} ({eppo_code})")
    print(f"📊 Regular questions: {len(answers)}")

    # Process regular questions
    print("\n" + "=" * 80)
    print("PROCESSING REGULAR QUESTIONS")
    print("=" * 80)

    for i, answer in enumerate(answers, 1):
        print(f"\n[{i}/{len(answers)}] {answer['code']}")

        existing = answer['existing_justification']
        if existing:
            print(f"📄 Found existing ({len(existing)} chars)")
            if skip_existing:
                print(f"⏭️  Skipped (existing justification)")
                # Record skipped question
                if track_costs and cost_tracker:
                    skip_metrics = QuestionMetrics(
                        species_name=species_name,
                        question_code=answer['code'],
                        question_text=answer['text'][:100],
                        status="skipped"
                    )
                    cost_tracker.record_question(skip_metrics)
                continue

        try:
            ai_text, metrics = await research_justification(
                species_name=species_name,
                question_code=answer['code'],
                question_text=answer['text'],
                question_info=answer['info'],
                exclude_domains=exclude_domains or [],
                track_metrics=track_costs,
                use_hybrid=use_hybrid
            )

            # Record metrics
            if track_costs and cost_tracker and metrics:
                cost_tracker.record_question(metrics)

            combined = f"{existing}\n\n{ai_text}" if existing else ai_text
            update_answer_justification(db_path, answer['idAnswer'], combined)

            print(f"✅ Updated ({len(combined)} chars)")
        except Exception as e:
            print(f"❌ Error: {str(e)}")
            # Record error
            if track_costs and cost_tracker:
                error_metrics = QuestionMetrics(
                    species_name=species_name,
                    question_code=answer['code'],
                    question_text=answer['text'][:100],
                    status="error",
                    error_message=str(e)[:200]
                )
                cost_tracker.record_question(error_metrics)

    # Process pathway questions
    if process_pathways:
        try:
            pathways = get_assessment_pathways(db_path, assessment_id)
        except Exception as e:
            print(f"\n⚠️  Error getting pathways: {e}")
            pathways = []

        if pathways:
            print(f"\n{'=' * 80}")
            print(f"PROCESSING PATHWAY QUESTIONS ({len(pathways)} pathways)")
            print(f"{'=' * 80}")

            try:
                pathway_questions = get_pathway_questions(db_path)
            except Exception as e:
                print(f"\n⚠️  Error getting pathway questions: {e}")
                if track_costs and cost_tracker:
                    cost_tracker.end_species()
                return

            # Filter pathway questions if question_filter is set
            if question_filter:
                filter_code = question_filter.upper().rstrip('.')
                pathway_questions = [pq for pq in pathway_questions
                                    if pq['code'].upper().rstrip('.') == filter_code]
                if not pathway_questions:
                    print(f"⚠️  No matching pathway question found for {filter_code}")

            total = len(pathways) * len(pathway_questions)
            count = 0

            for pathway in pathways:
                pathway_name = pathway['name']
                print(f"\n📍 Pathway: {pathway_name}")

                for pq in pathway_questions:
                    count += 1
                    print(f"\n[{count}/{total}] {pq['code']} for {pathway_name}")

                    existing = get_existing_pathway_justification(
                        db_path, pathway['idEntryPathway'], pq['idPathQuestion'])

                    if existing:
                        print(f"📄 Found existing ({len(existing)} chars)")
                        if skip_existing:
                            print(f"⏭️  Skipped (existing justification)")
                            # Record skipped
                            if track_costs and cost_tracker:
                                skip_metrics = QuestionMetrics(
                                    species_name=species_name,
                                    question_code=pq['code'],
                                    question_text=pq['text'][:100],
                                    pathway_name=pathway_name,
                                    status="skipped"
                                )
                                cost_tracker.record_question(skip_metrics)
                            continue

                    try:
                        ai_text, metrics = await research_justification(
                            species_name=species_name,
                            question_code=pq['code'],
                            question_text=pq['text'],
                            question_info=pq['info'],
                            pathway_name=pathway_name,
                            exclude_domains=exclude_domains or [],
                            track_metrics=track_costs,
                            use_hybrid=use_hybrid
                        )

                        # Record metrics
                        if track_costs and cost_tracker and metrics:
                            cost_tracker.record_question(metrics)

                        combined = f"{existing}\n\n{ai_text}" if existing else ai_text
                        update_pathway_justification(
                            db_path, pathway['idEntryPathway'],
                            pq['idPathQuestion'], combined)

                        print(f"✅ Updated ({len(combined)} chars)")
                    except Exception as e:
                        print(f"❌ Error: {str(e)}")
                        # Record error
                        if track_costs and cost_tracker:
                            error_metrics = QuestionMetrics(
                                species_name=species_name,
                                question_code=pq['code'],
                                question_text=pq['text'][:100],
                                pathway_name=pathway_name,
                                status="error",
                                error_message=str(e)[:200]
                            )
                            cost_tracker.record_question(error_metrics)
        else:
            print("\nℹ️  No pathways selected for this assessment")

    # End tracking this species
    if track_costs and cost_tracker:
        cost_tracker.end_species()

async def main(source_db: str = DEFAULT_DB_PATH,
               output_dir: str = DEFAULT_OUTPUT_DIR,
               assessment_id: int = None,
               limit_questions: int = None,
               exclude_domains: List[str] = None,
               process_pathways: bool = True,
               skip_existing: bool = None,
               species_filter: List[str] = None,
               track_costs: bool = None,
               add_all_pathways: bool = False,
               question_filter: str = None):
    """Main workflow.

    Args:
        add_all_pathways: If True, automatically add all available pathways to
            each assessment before processing pathway questions.
        question_filter: If set, only process this specific question code.
    """
    global cost_tracker

    # Use configuration value if not explicitly set via command line
    if skip_existing is None:
        skip_existing = SKIP_EXISTING_JUSTIFICATION
    if track_costs is None:
        track_costs = ENABLE_COST_TRACKING

    print("\n" + "=" * 80)
    print("BioPRIO JUSTIFICATION POPULATOR (HYBRID)")
    print("=" * 80)

    print(f"\n📂 Source Database: {source_db}")
    print(f"📂 Skip existing justifications: {skip_existing}")
    print(f"📊 Cost tracking: {'Enabled' if track_costs else 'Disabled'}")
    print(f"🔬 Research mode: HYBRID (web + local documents)")
    print(f"📂 Species docs path: {SPECIES_DOCS_BASE_PATH}")
    if add_all_pathways:
        print(f"🛤️  Auto-add all pathways: Enabled")

    if exclude_domains is None:
        exclude_domains = EXCLUDED_DOMAINS

    if exclude_domains:
        print(f"\n⛔ Excluded: {', '.join(exclude_domains)}")

    # Determine question filter to use (command-line overrides config)
    effective_question_filter = question_filter if question_filter else QUESTION_FILTER

    # Initialize cost tracker
    if track_costs:
        # SMART_LLM is in '<provider>:<model>' format; strip the provider for the price lookup.
        smart_llm_env = os.environ.get("SMART_LLM", "gpt-4o-mini")
        model_name = smart_llm_env.split(":", 1)[-1]
        cost_tracker = CostTracker(model_name=model_name)
        print(f"💰 Tracking costs for model: {model_name}")

    # Copy database
    working_db = copy_database(source_db, output_dir)

    print(f"\n✅ Working with: {working_db}")
    print(f"✅ Complete structure preserved")

    # Confirm (skip if filtering to single question or limited questions)
    if not effective_question_filter and (limit_questions is None or limit_questions > 5):
        response = input("\nThis will make many API calls. Continue? (yes/no): ")
        if response.lower() not in ['yes', 'y']:
            print("Cancelled.")
            return

    # Process
    print("\n" + "=" * 80)
    print("STARTING RESEARCH")
    print("=" * 80)
    if skip_existing:
        print("ℹ️  Skip existing justifications: Enabled")
    else:
        print("ℹ️  Existing justifications preserved, AI text appended")
    if process_pathways:
        print("ℹ️  Will process pathway questions for each selected pathway")
    if effective_question_filter:
        print(f"🔍 Question filter: {effective_question_filter.upper()} only")

    # Determine species filter to use (command-line overrides config)
    effective_filter = species_filter if species_filter else (SPECIES_FILTER if SPECIES_FILTER else None)

    # Get list of assessments to process
    if assessment_id:
        assessment_ids = [assessment_id]
        print(f"\nℹ️  Processing single assessment: {assessment_id}")
    elif effective_filter:
        assessment_ids = get_all_assessment_ids(working_db, effective_filter)
        print(f"\nℹ️  Filtering by species: {effective_filter}")
        print(f"    Found {len(assessment_ids)} matching assessment(s)")
        # Verify all requested filters were found
        if assessment_ids:
            found_ids = get_species_identifiers_for_assessments(working_db, assessment_ids)
            found_upper = set(f.upper() for f in found_ids)
            missing = [f for f in effective_filter if f.upper() not in found_upper]
            if missing:
                print(f"⚠️  Warning: No assessments found for: {missing}")
    else:
        assessment_ids = get_all_assessment_ids(working_db)
        print(f"\nℹ️  Processing all assessments: {len(assessment_ids)} total")

    # Process each assessment
    try:
        for idx, aid in enumerate(assessment_ids, 1):
            if len(assessment_ids) > 1:
                print("\n" + "=" * 80)
                print(f"ASSESSMENT {idx}/{len(assessment_ids)} (ID: {aid})")
                print("=" * 80)

            await process_assessment(
                db_path=working_db,
                assessment_id=aid,
                exclude_domains=exclude_domains,
                limit_questions=limit_questions,
                process_pathways=process_pathways,
                skip_existing=skip_existing,
                track_costs=track_costs,
                add_all_pathways=add_all_pathways,
                question_filter=effective_question_filter
            )
    finally:
        # Clean up temp documents folder
        cleanup_temp_docs()

    # Finalize cost tracking and export
    if track_costs and cost_tracker:
        cost_tracker.finalize()
        totals = cost_tracker.get_totals()

        print("\n" + "=" * 80)
        print("💰 COST SUMMARY")
        print("=" * 80)
        print(f"Species processed: {totals['total_species']}")
        print(f"Questions processed: {totals['total_questions_processed']}")
        print(f"Questions skipped: {totals['total_questions_skipped']}")
        print(f"Questions failed: {totals['total_questions_failed']}")
        print(f"Total tokens: ~{totals['total_tokens']:,}")
        print(f"Total searches: ~{totals['total_searches']}")
        print(f"Estimated LLM cost: ${totals['estimated_llm_cost']:.4f}")
        print(f"Estimated search cost: ${totals['estimated_search_cost']:.4f}")
        print(f"Estimated TOTAL cost: ${totals['estimated_total_cost']:.4f}")
        print(f"Total duration: {totals['total_duration_seconds']/60:.1f} minutes")

        # Export to Excel (prints saved location internally)
        cost_tracker.export_to_excel(output_dir)

    print("\n" + "=" * 80)
    print("✅ COMPLETED")
    print("=" * 80)
    print(f"\n📁 Database: {working_db}")
    print("\n✅ Regular questions: AI text appended to answers table")
    if process_pathways:
        print("✅ Pathway questions: AI text appended to pathwayAnswers table")
    if track_costs and cost_tracker and EXCEL_AVAILABLE:
        print(f"📊 Cost report: {output_dir}")
    print("\n🚀 Ready to use in BioPRIO app!")

# =============================================================================
# ENTRY POINT
# =============================================================================

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="BioPRIO Justification Populator")
    parser.add_argument('--db', type=str, default=DEFAULT_DB_PATH)
    parser.add_argument('--output', type=str, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument('--assessment-id', type=int, default=None)
    parser.add_argument('--limit-questions', type=int, default=None)
    parser.add_argument('--no-pathways', action='store_true',
                       help='Skip pathway questions')
    parser.add_argument('--species', type=str, nargs='+', default=None,
                       help='Filter by species (EPPO codes, names, or GBIF keys)')
    parser.add_argument('--overwrite', action='store_true',
                       help=f'Overwrite existing justifications (default: SKIP_EXISTING_JUSTIFICATION={SKIP_EXISTING_JUSTIFICATION})')
    parser.add_argument('--exclude-domains', type=str, nargs='+', default=None)
    parser.add_argument('--no-default-exclusions', action='store_true')
    parser.add_argument('--no-cost-tracking', action='store_true',
                       help='Disable cost tracking and Excel report generation')
    parser.add_argument('--cost-report-name', type=str, default=None,
                       help=f'Custom name for cost report Excel file (default: {COST_REPORT_FILENAME})')
    parser.add_argument('--add-all-pathways', action='store_true',
                       help='Automatically add all available pathways to each assessment before processing')
    parser.add_argument('--question', type=str, default=None,
                       help='Process only specific question code (e.g., --question EST2, --question ENT2A)')

    args = parser.parse_args()

    # Build exclusion list
    exclude_domains = None
    if not args.no_default_exclusions:
        exclude_domains = EXCLUDED_DOMAINS.copy()
        if args.exclude_domains:
            exclude_domains.extend(args.exclude_domains)
    elif args.exclude_domains:
        exclude_domains = args.exclude_domains

    # Determine skip_existing based on command line flag or use config default
    skip_existing = False if args.overwrite else None  # None means use config default

    # Determine cost tracking
    track_costs = False if args.no_cost_tracking else None  # None means use config default

    # Update cost report filename if specified
    if args.cost_report_name:
        COST_REPORT_FILENAME = args.cost_report_name

    asyncio.run(main(
        source_db=args.db,
        output_dir=args.output,
        assessment_id=args.assessment_id,
        limit_questions=args.limit_questions,
        exclude_domains=exclude_domains,
        process_pathways=not args.no_pathways,
        skip_existing=skip_existing,
        species_filter=args.species,
        track_costs=track_costs,
        add_all_pathways=args.add_all_pathways,
        question_filter=args.question
    ))
