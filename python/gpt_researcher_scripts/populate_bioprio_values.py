"""
Populate BioPRIO Min/Likely/Max Values from AI Justifications

Adapted from FinnPRIO for terrestrial invertebrates.

This script:
1. Reads the AI-enhanced database (output from populate_bioprio_justifications.py)
2. For each answer with justification, uses GPT-4o to determine appropriate min/likely/max values
3. Updates the database with selected option codes
4. Tracks API costs and exports report

Usage:
    python populate_bioprio_values.py
    python populate_bioprio_values.py --db path/to/database.db
    python populate_bioprio_values.py --assessment-id 5
    python populate_bioprio_values.py --species "Formica aserva"
    python populate_bioprio_values.py --question ENT1
"""

import sys
import sqlite3
import json
import os
import asyncio
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from dataclasses import dataclass, field
import argparse
from datetime import datetime
from openai import AsyncOpenAI

# Import instructions loader for Rmd-based value selection prompts
from bioprio_instructions_loader import build_value_selection_prompt, get_question_instructions

# DAG enforcement layer (modules live in parent python/ directory)
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from dag_values import (
    topological_sort_answers,
    check_zero_forcing,
    check_sibling_clamp,
    build_scored_prior_context,
    append_dag_correction,
)

# Check for openpyxl availability
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

################################################################################
# CONFIGURATION
################################################################################

SKIP_EXISTING_VALUES = False

# API Keys
OPENAI_API_KEY_FILE = r"C:\Users\dafl\OneDrive - Folkehelseinstituttet\API keys\tore_vkm_openai.txt"

def load_api_key(file_path: str) -> str:
    try:
        with open(file_path, 'r') as f:
            return f.read().strip()
    except FileNotFoundError:
        print(f"⚠️  Warning: API key file not found: {file_path}")
        return ""

os.environ['OPENAI_API_KEY'] = load_api_key(OPENAI_API_KEY_FILE)

# Database Path
INPUT_DATABASE = r"C:\Users\dafl\OneDrive - Folkehelseinstituttet\FinnPrio\BioPRIO_development\databases\ants_ai\ants_Minimal_ai_ai_enhanced_28_04_2026.db"

# Filter by species identifiers (empty list = process all species)
# Supports: EPPO codes, scientific names, or GBIF taxon keys
# "1315155", "1317433"
SPECIES_FILTER = []

# Filter by question code (None = process all questions)
# Examples: "ENT1", "EST4", "IMP2", "ENT2A"
QUESTION_FILTER = None

# Cost Tracking Configuration
ENABLE_COST_TRACKING = True
COST_REPORT_FILENAME = "values_cost_report"

# OpenAI pricing (per 1M tokens)
OPENAI_PRICING = {
    "gpt-4o-mini": {"input": 0.15, "output": 0.60},
    "gpt-4o": {"input": 2.50, "output": 10.00},
    "gpt-4-turbo": {"input": 10.00, "output": 30.00},
    "gpt-3.5-turbo": {"input": 0.50, "output": 1.50},
}

################################################################################
# COST TRACKING
################################################################################

@dataclass
class QuestionMetrics:
    """Metrics for a single question API call."""
    species_name: str
    question_code: str
    start_time: Optional[datetime] = None
    end_time: Optional[datetime] = None
    duration_seconds: float = 0.0
    input_tokens: int = 0
    output_tokens: int = 0
    total_tokens: int = 0
    estimated_cost: float = 0.0
    status: str = "pending"  # pending, success, error, skipped
    error_message: str = ""


@dataclass
class SpeciesMetrics:
    """Aggregated metrics for a species."""
    species_name: str
    identifier: str
    assessment_id: int
    start_time: Optional[datetime] = None
    end_time: Optional[datetime] = None
    total_duration_seconds: float = 0.0
    questions_processed: int = 0
    questions_skipped: int = 0
    questions_failed: int = 0
    total_input_tokens: int = 0
    total_output_tokens: int = 0
    total_tokens: int = 0
    estimated_cost: float = 0.0
    question_metrics: List[QuestionMetrics] = field(default_factory=list)


class CostTracker:
    """Tracks costs and metrics across the entire run."""

    def __init__(self, model_name: str = "gpt-4o-mini"):
        self.model_name = model_name
        self.species_metrics: List[SpeciesMetrics] = []
        self.current_species: Optional[SpeciesMetrics] = None
        self.run_start_time: datetime = datetime.now()
        self.run_end_time: Optional[datetime] = None
        self.input_price_per_1m = OPENAI_PRICING.get(model_name, {}).get("input", 0.15)
        self.output_price_per_1m = OPENAI_PRICING.get(model_name, {}).get("output", 0.60)

    def start_species(self, species_name: str, identifier: str, assessment_id: int):
        self.current_species = SpeciesMetrics(
            species_name=species_name,
            identifier=identifier,
            assessment_id=assessment_id,
            start_time=datetime.now()
        )

    def end_species(self):
        if self.current_species:
            self.current_species.end_time = datetime.now()
            self.current_species.total_duration_seconds = (
                self.current_species.end_time - self.current_species.start_time
            ).total_seconds()
            for qm in self.current_species.question_metrics:
                self.current_species.total_input_tokens += qm.input_tokens
                self.current_species.total_output_tokens += qm.output_tokens
                self.current_species.total_tokens += qm.total_tokens
                self.current_species.estimated_cost += qm.estimated_cost
            self.species_metrics.append(self.current_species)
            self.current_species = None

    def record_question(self, metrics: QuestionMetrics):
        if self.current_species:
            metrics.estimated_cost = (
                (metrics.input_tokens * self.input_price_per_1m / 1_000_000) +
                (metrics.output_tokens * self.output_price_per_1m / 1_000_000)
            )
            self.current_species.question_metrics.append(metrics)
            if metrics.status == "success":
                self.current_species.questions_processed += 1
            elif metrics.status == "skipped":
                self.current_species.questions_skipped += 1
            elif metrics.status == "error":
                self.current_species.questions_failed += 1

    def finalize(self):
        self.run_end_time = datetime.now()

    def get_totals(self) -> Dict:
        return {
            "total_species": len(self.species_metrics),
            "total_duration_seconds": sum(s.total_duration_seconds for s in self.species_metrics),
            "total_questions_processed": sum(s.questions_processed for s in self.species_metrics),
            "total_questions_skipped": sum(s.questions_skipped for s in self.species_metrics),
            "total_questions_failed": sum(s.questions_failed for s in self.species_metrics),
            "total_input_tokens": sum(s.total_input_tokens for s in self.species_metrics),
            "total_output_tokens": sum(s.total_output_tokens for s in self.species_metrics),
            "total_tokens": sum(s.total_tokens for s in self.species_metrics),
            "estimated_cost": sum(s.estimated_cost for s in self.species_metrics),
        }

    def export_to_excel(self, output_dir: str, filename: str = None):
        if not EXCEL_AVAILABLE:
            print("⚠️  Cannot export to Excel: openpyxl not installed")
            return None
        if not self.species_metrics:
            print("⚠️  No metrics to export")
            return None

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = filename or COST_REPORT_FILENAME
        filepath = Path(output_dir) / f"{filename}_{timestamp}.xlsx"

        wb = openpyxl.Workbook()
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))

        # Summary sheet
        ws = wb.active
        ws.title = "Summary"
        totals = self.get_totals()
        run_duration = (self.run_end_time - self.run_start_time).total_seconds() if self.run_end_time else 0

        summary = [
            ["BioPRIO Values Cost Report", ""],
            ["", ""],
            ["Run Information", ""],
            ["Start Time", self.run_start_time.strftime("%Y-%m-%d %H:%M:%S")],
            ["End Time", self.run_end_time.strftime("%Y-%m-%d %H:%M:%S") if self.run_end_time else "N/A"],
            ["Duration", f"{run_duration/60:.1f} minutes"],
            ["Model", self.model_name],
            ["", ""],
            ["Summary", ""],
            ["Species Processed", totals["total_species"]],
            ["Questions Processed", totals["total_questions_processed"]],
            ["Questions Skipped", totals["total_questions_skipped"]],
            ["Questions Failed", totals["total_questions_failed"]],
            ["", ""],
            ["Tokens", ""],
            ["Input Tokens", totals["total_input_tokens"]],
            ["Output Tokens", totals["total_output_tokens"]],
            ["Total Tokens", totals["total_tokens"]],
            ["", ""],
            ["Cost", ""],
            ["Estimated Total", f"${totals['estimated_cost']:.4f}"],
        ]
        for row_idx, row in enumerate(summary, 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = Font(bold=True, size=14)
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25

        # Per species sheet
        ws2 = wb.create_sheet("Per Species")
        headers = ["Species", "ID", "Assessment", "Duration (min)", "Questions", "Skipped", "Failed",
                   "Input Tokens", "Output Tokens", "Cost ($)"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        for row_idx, sm in enumerate(self.species_metrics, 2):
            data = [sm.species_name, sm.identifier, sm.assessment_id,
                    sm.total_duration_seconds / 60, sm.questions_processed,
                    sm.questions_skipped, sm.questions_failed,
                    sm.total_input_tokens, sm.total_output_tokens, sm.estimated_cost]
            for col_idx, v in enumerate(data, 1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=v)
                cell.border = border

        wb.save(filepath)
        print(f"\n📊 Cost report saved: {filepath}")
        return str(filepath)


# Global instance
cost_tracker: Optional[CostTracker] = None

################################################################################
# MAIN CLASS
################################################################################

client = AsyncOpenAI(api_key=os.environ.get('OPENAI_API_KEY', ''))

class ValuePopulator:
    def __init__(self, db_path: str, assessment_id: Optional[int] = None):
        self.db_path = db_path
        self.assessment_id = assessment_id
        self.conn = None

    def connect(self):
        self.conn = sqlite3.connect(self.db_path)
        self.conn.row_factory = sqlite3.Row

    def disconnect(self):
        if self.conn:
            self.conn.close()

    def get_all_assessment_ids(self, species_filter: List[str] = None) -> List[int]:
        """Get assessment IDs, optionally filtered by species identifiers."""
        cursor = self.conn.cursor()
        if species_filter:
            placeholders = ','.join(['?' for _ in species_filter])
            upper_filters = [f.upper() for f in species_filter]
            cursor.execute(f"""
                SELECT a.idAssessment
                FROM assessments a
                JOIN pests p ON a.idPest = p.idPest
                WHERE UPPER(p.eppoCode) IN ({placeholders})
                   OR UPPER(p.scientificName) IN ({placeholders})
                   OR p.gbifTaxonKey IN ({placeholders})
                ORDER BY a.idAssessment
            """, upper_filters + upper_filters + species_filter)
        else:
            cursor.execute("SELECT idAssessment FROM assessments ORDER BY idAssessment")
        return [row['idAssessment'] for row in cursor.fetchall()]

    def get_species_identifiers_for_assessments(self, assessment_ids: List[int]) -> List[str]:
        if not assessment_ids:
            return []
        cursor = self.conn.cursor()
        placeholders = ','.join(['?' for _ in assessment_ids])
        cursor.execute(f"""
            SELECT DISTINCT COALESCE(NULLIF(p.eppoCode, ''), p.scientificName) as identifier
            FROM assessments a JOIN pests p ON a.idPest = p.idPest
            WHERE a.idAssessment IN ({placeholders})
        """, assessment_ids)
        return [row['identifier'] for row in cursor.fetchall() if row['identifier']]

    def get_question_options(self, id_question: int, table: str = "questions") -> Dict:
        """
        Get question details and options from database.

        Returns:
            Dict with 'question', 'options', 'type', 'code', and 'info' keys
        """
        cursor = self.conn.cursor()
        if table == "questions":
            # Include group, number, subgroup to build question code
            cursor.execute(
                'SELECT question, list, type, info, "group", number, subgroup FROM questions WHERE idQuestion = ?',
                (id_question,)
            )
        else:
            # Pathway questions have group and number
            cursor.execute(
                'SELECT question, list, info, "group", number FROM pathwayQuestions WHERE idPathQuestion = ?',
                (id_question,)
            )
        row = cursor.fetchone()
        if not row:
            return None

        # Build question code (e.g., "ENT1", "EST4", "IMP2.1", "ENT2A")
        if table == "questions":
            question_type = row['type'] if row['type'] else "minmax"
            group = row['group'] or ""
            number = row['number'] or ""
            subgroup = row['subgroup'] or ""
            if subgroup:
                question_code = f"{group}{number}.{subgroup}"
            else:
                question_code = f"{group}{number}"
        else:
            question_type = "minmax"
            group = row['group'] or ""
            number = row['number'] or ""
            question_code = f"{group}{number}"

        return {
            'question': row['question'],
            'options': json.loads(row['list']),
            'type': question_type,
            'code': question_code,
            'info': row['info'] if row['info'] else ""
        }

    def get_species_name(self, id_assessment: int) -> str:
        cursor = self.conn.cursor()
        cursor.execute("""
            SELECT p.scientificName FROM assessments a
            JOIN pests p ON a.idPest = p.idPest WHERE a.idAssessment = ?
        """, (id_assessment,))
        row = cursor.fetchone()
        return row['scientificName'] if row else "Unknown"

    def get_species_identifier(self, id_assessment: int) -> str:
        cursor = self.conn.cursor()
        cursor.execute("""
            SELECT COALESCE(NULLIF(p.eppoCode, ''), p.gbifTaxonKey, '') as identifier
            FROM assessments a JOIN pests p ON a.idPest = p.idPest WHERE a.idAssessment = ?
        """, (id_assessment,))
        row = cursor.fetchone()
        return row['identifier'] if row else ""

    async def _call_gpt_boolean(
        self, justification: str, question_code: str, yes_code: str
    ) -> Tuple[Optional[Dict], int, int]:
        """Ask a simple yes/no question for boolean sub-questions (IMP2.x, IMP4.x).

        Returns ({min: yes_code, likely: yes_code, max: yes_code}, in_tok, out_tok) for YES,
                ({min: None, likely: None, max: None}, in_tok, out_tok) for NO,
                (None, 0, 0) on error.
        """
        guidance = []
        try:
            q = get_question_instructions(question_code)
            question_text = f"{q['code']}: {q['text']}"
            guidance = q.get('guidance', [])
        except Exception:
            question_text = question_code
            guidance = []

        guidance_block = (
            "\nGUIDANCE:\n" + "\n".join(f"- {g}" for g in guidance)
        ) if guidance else ""

        prompt = (
            f"Does the following justification indicate that this applies to the species?\n\n"
            f"QUESTION: {question_text}"
            f"{guidance_block}\n\n"
            f"JUSTIFICATION:\n{justification}\n\n"
            f"Answer YES if the justification supports it, "
            f"NO if it does not occur or is not mentioned.\n"
            f'Return ONLY: {{"answer": "YES"}} or {{"answer": "NO"}}'
        )

        try:
            response = await client.chat.completions.create(
                model=os.getenv("LLM_MODEL", "gpt-4o"),
                messages=[
                    {"role": "system", "content": "You are an expert in invasive species/arthropod risk assessment."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.0,
                max_tokens=20
            )
            content = response.choices[0].message.content.strip()
            in_tok = response.usage.prompt_tokens if response.usage else 0
            out_tok = response.usage.completion_tokens if response.usage else 0
            if "```" in content:
                content = content.split("```")[1].split("```")[0].strip()
                if content.startswith("json"):
                    content = content[4:].strip()
            result = json.loads(content)
            if result.get("answer", "").upper() == "YES":
                return {"min": yes_code, "likely": yes_code, "max": yes_code}, in_tok, out_tok
            return {"min": None, "likely": None, "max": None}, in_tok, out_tok
        except Exception as e:
            print(f"  ⚠️  Error in boolean evaluation: {type(e).__name__}: {e}")
            return None, 0, 0

    async def determine_values_with_gpt(
        self,
        species_name: str,
        question_text: str,
        options: List[Dict],
        justification: str,
        question_type: str = "minmax",
        question_code: str = None,
        prior_context: str = "",
    ) -> Tuple[Optional[Dict], int, int]:
        """
        Use GPT to determine appropriate min/likely/max values based on justification.

        Uses Rmd-based instructions via build_value_selection_prompt() for comprehensive
        prompts with options, guidance, and scoring criteria.

        Args:
            species_name: Scientific name of the species
            question_text: The question text (unused, kept for compatibility)
            options: List of option dicts with 'opt', 'text', 'points'
            justification: The AI-generated justification to analyze
            question_type: 'minmax' or 'boolean'
            question_code: Question code (e.g., 'ENT1', 'EST4') - required for Rmd lookup
            prior_context: Scored upstream values to inject into the prompt (may be empty)

        Returns:
            Tuple of (values_dict, input_tokens, output_tokens) or (None, 0, 0) on error
        """
        # Boolean sub-questions (IMP2.x, IMP4.x): route through yes/no path
        if question_type == 'boolean' and options:
            yes_code = options[0]['opt']
            return await self._call_gpt_boolean(justification, question_code, yes_code)

        prompt = build_value_selection_prompt(question_code, species_name, justification, options)
        if prior_context:
            prompt = prior_context + "\n\n" + prompt
        return await self._call_gpt_for_values(prompt, options)

    async def _call_gpt_for_values(
        self,
        prompt: str,
        options: List[Dict]
    ) -> Tuple[Optional[Dict], int, int]:
        """
        Call GPT API with prompt and parse response.

        Returns:
            Tuple of (values_dict, input_tokens, output_tokens) or (None, 0, 0) on error
        """
        try:
            response = await client.chat.completions.create(
                model=os.getenv("LLM_MODEL", "gpt-4o"),
                messages=[
                    {"role": "system", "content": "You are an expert in invasive species risk assessment. You analyze scientific evidence and determine appropriate risk estimates. Return only JSON."},
                    {"role": "user", "content": prompt}
                ],
                temperature=float(os.getenv("TEMPERATURE", "0.1")),
                max_tokens=int(os.getenv("LLM_MAX_TOKENS", "500"))
            )

            content = response.choices[0].message.content.strip()
            input_tokens = response.usage.prompt_tokens if response.usage else 0
            output_tokens = response.usage.completion_tokens if response.usage else 0

            # Extract JSON from response (in case model adds extra text)
            if "```json" in content:
                content = content.split("```json")[1].split("```")[0].strip()
            elif "```" in content:
                content = content.split("```")[1].split("```")[0].strip()

            # Parse JSON
            values = json.loads(content)

            # Build mapping from points to option codes (for when GPT returns integers)
            points_to_opt = {opt['points']: opt['opt'] for opt in options}
            valid_opts = {opt['opt'] for opt in options}

            # Convert values to lowercase and handle integers
            for key in ['min', 'likely', 'max']:
                if key in values and values[key] is not None:
                    val = values[key]
                    # Handle integer responses (GPT sometimes returns points instead of letters)
                    if isinstance(val, int):
                        if val in points_to_opt:
                            values[key] = points_to_opt[val]
                        else:
                            # Try to convert 1->a, 2->b, etc.
                            values[key] = chr(ord('a') + val - 1) if 1 <= val <= 26 else str(val)
                    else:
                        values[key] = str(val).lower()

            # Validate that all keys exist and values are valid option codes
            required_keys = ['min', 'likely', 'max']
            if not all(k in values for k in required_keys):
                raise ValueError(f"Missing required keys. Got: {values.keys()}")

            for key in required_keys:
                # Allow None/null for boolean questions (when answer is NO)
                if values[key] is not None and values[key] not in valid_opts:
                    raise ValueError(f"Invalid option code '{values[key]}' for {key}. Valid options: {valid_opts}")

            return values, input_tokens, output_tokens

        except json.JSONDecodeError as e:
            print(f"  ⚠️  JSON parsing error: {e}")
            print(f"  Response content: {content if 'content' in locals() else 'N/A'}")
            return None, 0, 0
        except Exception as e:
            print(f"  ⚠️  Error determining values: {type(e).__name__}: {e}")
            return None, 0, 0

    def update_answer_values(self, id_answer: int, min_val: str, likely_val: str, max_val: str):
        cursor = self.conn.cursor()
        cursor.execute("UPDATE answers SET min=?, likely=?, max=? WHERE idAnswer=?",
                      (min_val, likely_val, max_val, id_answer))
        self.conn.commit()

    def update_pathway_answer_values(self, id_path_answer: int, min_val: str, likely_val: str, max_val: str):
        cursor = self.conn.cursor()
        cursor.execute("UPDATE pathwayAnswers SET min=?, likely=?, max=? WHERE idPathAnswer=?",
                      (min_val, likely_val, max_val, id_path_answer))
        self.conn.commit()

    def get_answers_to_populate(self) -> List[Dict]:
        cursor = self.conn.cursor()
        where = "AND a.idAssessment = ?" if self.assessment_id else ""
        params = [self.assessment_id] if self.assessment_id else []
        cursor.execute(f"""
            SELECT a.idAnswer, a.idAssessment, a.idQuestion, a.justification, a.min, a.likely, a.max
            FROM answers a WHERE a.justification IS NOT NULL AND a.justification != '' {where}
        """, params)
        return [{
            'idAnswer': r['idAnswer'], 'idAssessment': r['idAssessment'],
            'idQuestion': r['idQuestion'], 'justification': r['justification'],
            'has_values': bool(r['min'] and r['likely'] and r['max'])
        } for r in cursor.fetchall()]

    def get_pathway_answers_to_populate(self) -> List[Dict]:
        cursor = self.conn.cursor()
        where = "AND ep.idAssessment = ?" if self.assessment_id else ""
        params = [self.assessment_id] if self.assessment_id else []
        # Use CASE to properly map idPathQuestion to codes (A/B suffix not in DB)
        # idPathQuestion 1=ENT2A, 2=ENT2B, 3=ENT3, 4=ENT4
        cursor.execute(f"""
            SELECT pa.idPathAnswer, ep.idAssessment, pa.idPathQuestion, pa.justification,
                   pa.min, pa.likely, pa.max, pa.idEntryPathway, p.name as pathway_name,
                   CASE pa.idPathQuestion
                       WHEN 1 THEN 'ENT2A'
                       WHEN 2 THEN 'ENT2B'
                       WHEN 3 THEN 'ENT3'
                       WHEN 4 THEN 'ENT4'
                       ELSE pq."group" || pq.number
                   END as question_code
            FROM pathwayAnswers pa
            JOIN entryPathways ep ON pa.idEntryPathway = ep.idEntryPathway
            JOIN pathways p ON ep.idPathway = p.idPathway
            JOIN pathwayQuestions pq ON pa.idPathQuestion = pq.idPathQuestion
            WHERE pa.justification IS NOT NULL AND pa.justification != '' {where}
        """, params)
        return [{
            'idPathAnswer': r['idPathAnswer'], 'idAssessment': r['idAssessment'],
            'idPathQuestion': r['idPathQuestion'], 'idEntryPathway': r['idEntryPathway'],
            'justification': r['justification'],
            'has_values': bool(r['min'] and r['likely'] and r['max']),
            'pathway_name': r['pathway_name'],
            'question_code': r['question_code']
        } for r in cursor.fetchall()]

    def get_max_option_for_question(self, id_question: int, table: str = "pathwayQuestions") -> str:
        """Get the maximum (highest points) option code for a question."""
        cursor = self.conn.cursor()
        if table == "pathwayQuestions":
            cursor.execute("SELECT list FROM pathwayQuestions WHERE idPathQuestion = ?", (id_question,))
        else:
            cursor.execute("SELECT list FROM questions WHERE idQuestion = ?", (id_question,))
        row = cursor.fetchone()
        if not row:
            return None
        options = json.loads(row['list'])
        # Find option with highest points
        max_opt = max(options, key=lambda x: x.get('points', 0))
        return max_opt['opt']

    async def populate_values_for_assessment(self, assessment_id: int, skip_existing: bool = True,
                                             track_costs: bool = True, question_filter: str = None):
        global cost_tracker
        original_id = self.assessment_id
        self.assessment_id = assessment_id

        try:
            species_name = self.get_species_name(assessment_id)
            species_id = self.get_species_identifier(assessment_id)
            print(f"\n{'='*60}\nAssessment {assessment_id}: {species_name}\n{'='*60}")

            if track_costs and cost_tracker:
                cost_tracker.start_species(species_name, species_id, assessment_id)

            answers = self.get_answers_to_populate()
            pathway_answers = self.get_pathway_answers_to_populate()

            if skip_existing:
                answers = [a for a in answers if not a['has_values']]
                pathway_answers = [p for p in pathway_answers if not p['has_values']]

            # Apply question filter if specified
            if question_filter:
                filter_upper = question_filter.upper()
                # Filter regular answers by looking up question code
                filtered_answers = []
                for ans in answers:
                    q_data = self.get_question_options(ans['idQuestion'], "questions")
                    if q_data and q_data['code'].upper() == filter_upper:
                        filtered_answers.append(ans)
                answers = filtered_answers

                # Filter pathway answers - check if filter matches pathway question codes
                if filter_upper.startswith("ENT") and any(c.isalpha() and c != 'E' and c != 'N' and c != 'T'
                                                          for c in filter_upper):
                    # Pathway question filter (ENT2A, ENT2B, ENT3, ENT4)
                    filtered_pathway = []
                    for ans in pathway_answers:
                        q_data = self.get_question_options(ans['idPathQuestion'], "pathwayQuestions")
                        if q_data and q_data['code'].upper() == filter_upper:
                            filtered_pathway.append(ans)
                    pathway_answers = filtered_pathway
                else:
                    # Regular question filter - clear pathway answers
                    pathway_answers = []

                print(f"Filtering by question: {question_filter}")

            total = len(answers) + len(pathway_answers)
            print(f"Regular: {len(answers)}, Pathway: {len(pathway_answers)}, Total: {total}")

            if total == 0:
                print("✅ Nothing to populate")
                if track_costs and cost_tracker:
                    cost_tracker.end_species()
                return 0

            # Process regular answers
            for i, ans in enumerate(answers, 1):
                q_data = self.get_question_options(ans['idQuestion'], "questions")
                if not q_data:
                    continue

                q_code = q_data['code']  # Use proper question code (ENT1, EST4, etc.)
                print(f"[{i}/{len(answers)}] {q_code}...", end=" ")

                start = datetime.now()
                values, in_tok, out_tok = await self.determine_values_with_gpt(
                    species_name, q_data['question'], q_data['options'],
                    ans['justification'], q_data['type'], question_code=q_code
                )
                duration = (datetime.now() - start).total_seconds()

                metrics = QuestionMetrics(
                    species_name=species_name, question_code=q_code,
                    start_time=start, end_time=datetime.now(),
                    duration_seconds=duration, input_tokens=in_tok, output_tokens=out_tok,
                    total_tokens=in_tok + out_tok
                )

                if values and not all(v is None for v in [values['min'], values['likely'], values['max']]):
                    self.update_answer_values(ans['idAnswer'], values['min'], values['likely'], values['max'])
                    print(f"✅ {values['min']}/{values['likely']}/{values['max']}")
                    metrics.status = "success"
                else:
                    print("⏭️ skipped")
                    metrics.status = "skipped"

                if track_costs and cost_tracker:
                    cost_tracker.record_question(metrics)

            # Process pathway answers
            for i, ans in enumerate(pathway_answers, 1):
                q_data = self.get_question_options(ans['idPathQuestion'], "pathwayQuestions")
                if not q_data:
                    continue

                q_code = q_data['code']  # Use proper question code (ENT2A, ENT2B, etc.)
                pathway_name = ans.get('pathway_name', '')
                print(f"[{i}/{len(pathway_answers)}] {q_code} ({pathway_name})...", end=" ")

                # Special handling for "Intentional introduction" pathway
                # ENT2A, ENT2B, ENT3 should all be set to MAXIMUM values
                is_intentional = pathway_name.lower() == "intentional introduction"
                is_entry_question = q_code in ["ENT2A", "ENT2B", "ENT3"]

                if is_intentional and is_entry_question:
                    # Get maximum option and set all values to max
                    max_opt = self.get_max_option_for_question(ans['idPathQuestion'], "pathwayQuestions")
                    if max_opt:
                        self.update_pathway_answer_values(ans['idPathAnswer'], max_opt, max_opt, max_opt)
                        print(f"✅ {max_opt}/{max_opt}/{max_opt} (INTENTIONAL: auto-max)")

                        metrics = QuestionMetrics(
                            species_name=species_name, question_code=q_code,
                            start_time=datetime.now(), end_time=datetime.now(),
                            duration_seconds=0, input_tokens=0, output_tokens=0, total_tokens=0,
                            status="success"
                        )
                        if track_costs and cost_tracker:
                            cost_tracker.record_question(metrics)
                        continue

                start = datetime.now()
                values, in_tok, out_tok = await self.determine_values_with_gpt(
                    species_name, q_data['question'], q_data['options'],
                    ans['justification'], q_data['type'], question_code=q_code
                )
                duration = (datetime.now() - start).total_seconds()

                metrics = QuestionMetrics(
                    species_name=species_name, question_code=q_code,
                    start_time=start, end_time=datetime.now(),
                    duration_seconds=duration, input_tokens=in_tok, output_tokens=out_tok,
                    total_tokens=in_tok + out_tok
                )

                if values and not all(v is None for v in [values['min'], values['likely'], values['max']]):
                    self.update_pathway_answer_values(ans['idPathAnswer'], values['min'], values['likely'], values['max'])
                    print(f"✅ {values['min']}/{values['likely']}/{values['max']}")
                    metrics.status = "success"
                else:
                    print("⏭️ skipped")
                    metrics.status = "skipped"

                if track_costs and cost_tracker:
                    cost_tracker.record_question(metrics)

            if track_costs and cost_tracker:
                cost_tracker.end_species()

            return total

        finally:
            self.assessment_id = original_id

    async def populate_values(self, skip_existing: bool = True, species_filter: List[str] = None,
                              question_filter: str = None, track_costs: bool = True):
        global cost_tracker

        print("\n" + "="*60)
        print("BioPRIO Value Populator")
        print("="*60)
        print(f"Database: {self.db_path}")
        print(f"Skip existing: {skip_existing}")
        print(f"Question filter: {question_filter or 'All questions'}")
        print(f"Cost tracking: {track_costs}")

        if track_costs:
            model = os.getenv("LLM_MODEL", "gpt-4o-mini")
            cost_tracker = CostTracker(model_name=model)
            print(f"Model: {model}")

        self.connect()

        try:
            effective_filter = species_filter if species_filter else (SPECIES_FILTER if SPECIES_FILTER else None)

            if self.assessment_id:
                assessment_ids = [self.assessment_id]
                print(f"\nProcessing single assessment: {self.assessment_id}")
            elif effective_filter:
                assessment_ids = self.get_all_assessment_ids(effective_filter)
                print(f"\nFiltering by: {effective_filter}")
                print(f"Found {len(assessment_ids)} assessment(s)")
            else:
                assessment_ids = self.get_all_assessment_ids()
                print(f"\nProcessing all: {len(assessment_ids)} assessments")

            total_processed = 0
            for idx, aid in enumerate(assessment_ids, 1):
                processed = await self.populate_values_for_assessment(
                    aid, skip_existing, track_costs, question_filter
                )
                total_processed += processed if processed else 0

            # Finalize cost tracking
            if track_costs and cost_tracker:
                cost_tracker.finalize()
                totals = cost_tracker.get_totals()
                print("\n" + "="*60)
                print("COST SUMMARY")
                print("="*60)
                print(f"Questions processed: {totals['total_questions_processed']}")
                print(f"Total tokens: {totals['total_tokens']}")
                print(f"Estimated cost: ${totals['estimated_cost']:.4f}")

                if EXCEL_AVAILABLE:
                    output_dir = Path(self.db_path).parent
                    cost_tracker.export_to_excel(str(output_dir))

            print("\n" + "="*60)
            print(f"✅ Done! Processed {total_processed} answers")
            print("="*60)

        finally:
            self.disconnect()


async def main(db_path: str = None, assessment_id: int = None,
               skip_existing: bool = None, species_filter: List[str] = None,
               question_filter: str = None, track_costs: bool = None):

    if skip_existing is None:
        skip_existing = SKIP_EXISTING_VALUES
    if track_costs is None:
        track_costs = ENABLE_COST_TRACKING

    if not db_path:
        db_path = INPUT_DATABASE

    if not db_path:
        # Auto-detect
        ai_dir = Path(__file__).parent.parent / "databases" / "ant_test"
        if ai_dir.exists():
            dbs = list(ai_dir.glob("*_ai_enhanced_*.db"))
            if dbs:
                db_path = str(sorted(dbs, key=lambda p: p.stat().st_mtime, reverse=True)[0])
                print(f"Auto-detected: {db_path}")

    if not db_path or not Path(db_path).exists():
        print(f"❌ Database not found: {db_path}")
        return

    populator = ValuePopulator(db_path, assessment_id)
    await populator.populate_values(skip_existing=skip_existing, species_filter=species_filter,
                                    question_filter=question_filter, track_costs=track_costs)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Populate min/likely/max values in BioPRIO database")
    parser.add_argument('--db', type=str, help="Path to database")
    parser.add_argument('--assessment-id', type=int, help="Process single assessment")
    parser.add_argument('--species', type=str, nargs='+', help="Filter by species (names, EPPO codes, or GBIF keys)")
    parser.add_argument('--question', type=str, help="Filter by question code (e.g., ENT1, EST4, ENT2A)")
    parser.add_argument('--overwrite', action='store_true', help="Overwrite existing values")
    parser.add_argument('--no-cost-tracking', action='store_true', help="Disable cost tracking")

    args = parser.parse_args()

    if not os.getenv("OPENAI_API_KEY"):
        print("❌ OPENAI_API_KEY not set")
        exit(1)

    asyncio.run(main(
        db_path=args.db,
        assessment_id=args.assessment_id,
        skip_existing=False if args.overwrite else None,
        species_filter=args.species,
        question_filter=args.question or QUESTION_FILTER,
        track_costs=False if args.no_cost_tracking else None
    ))
